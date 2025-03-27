"""
Excel出力モジュール

このモジュールでは、処理結果をExcel形式で出力するための機能を提供します。
Excel生成の基本機能、カスタムヘッダー設定、データ変換、スタイル適用などの機能が含まれます。
"""

import os
import logging
import tempfile
from pathlib import Path
from typing import List, Dict, Any, Optional, BinaryIO, Union
from datetime import datetime

import openpyxl
# スタイル関連のインポートを追加
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ..data.models import ProcessResult, ExcelConfig
from ..data.interfaces import ExcelExporterProtocol, ProcessResultProtocol
from ..utils.errors import ExcelExportError, with_error_handling


class ExcelExporter(ExcelExporterProtocol):
    """
    Excel出力クラス
    
    処理結果をExcel形式で出力します。
    Excel生成の基本機能、カスタムヘッダー設定、データ変換、スタイル適用などの機能が含まれます。
    """
    
    def __init__(self, config: ExcelConfig, filename_mapping: Dict[str, str] = None):
        """
        初期化
        
        Args:
            config: Excel出力設定
            filename_mapping: ファイル名のマッピング辞書（オプション）
        """
        self.logger = logging.getLogger(__name__)
        self.config = config
        self.filename_mapping = filename_mapping or {}
    
    @with_error_handling(ExcelExportError, "Excel出力処理でエラーが発生しました")
    def export(self, results: List[ProcessResultProtocol], output_path: Union[str, Path]) -> Path:
        """
        処理結果をExcelファイルに出力します。
        
        Args:
            results: 処理結果のリスト
            output_path: 出力ファイルのパス
            
        Returns:
            エクスポートされたファイルのパス
            
        Raises:
            ExcelExportError: Excel出力処理でエラーが発生した場合
        """
        self.logger.info(f"Excel出力開始: 結果数={len(results)}, 出力先={output_path}")
        
        # 文字列パスをPathオブジェクトに変換
        if isinstance(output_path, str):
            output_path = Path(output_path)
        
        # 出力ディレクトリが存在しない場合は作成
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 既存ファイルのバックアップ
        if output_path.exists():
            backup_path = self._create_backup(output_path)
            self.logger.info(f"既存ファイルをバックアップしました: {backup_path}")
        
        # 新しいワークブックを作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "スタイルタイトル"
        
        # ヘッダーの設定
        self._set_headers(sheet)
        
        # データの追加
        self._add_data(sheet, results)
        
        # 列幅の自動調整
        self._adjust_column_widths(sheet)
        
        # 保存
        try:
            workbook.save(output_path)
            self.logger.info(f"Excelファイルを保存しました: {output_path}")
            return output_path
        except Exception as e:
            self.logger.error(f"Excelファイルの保存エラー: {e}")
            raise ExcelExportError(f"Excelファイルの保存に失敗しました: {str(e)}", 
                                 output_path=str(output_path)) from e
    
    @with_error_handling(ExcelExportError, "Excelバイナリデータの生成でエラーが発生しました")
    def get_binary_data(self, results: List[ProcessResultProtocol]) -> bytes:
        """
        処理結果のExcelバイナリデータを取得します。
        
        Args:
            results: 処理結果のリスト
            
        Returns:
            Excelバイナリデータ
            
        Raises:
            ExcelExportError: Excel出力処理でエラーが発生した場合
        """
        self.logger.info(f"Excelバイナリデータ生成開始: 結果数={len(results)}")
        
        # 新しいワークブックを作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "スタイルタイトル"
        
        # ヘッダーの設定
        self._set_headers(sheet)
        
        # データの追加
        self._add_data(sheet, results)
        
        # 列幅の自動調整
        self._adjust_column_widths(sheet)
        
        # 一時ファイルに保存してバイナリデータを取得
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            
            workbook.save(tmp_path)
            
            with open(tmp_path, 'rb') as f:
                binary_data = f.read()
            
            # 一時ファイルを削除
            os.unlink(tmp_path)
            
            self.logger.info(f"Excelバイナリデータを生成しました: {len(binary_data)} バイト")
            return binary_data
            
        except Exception as e:
            self.logger.error(f"Excelバイナリデータの生成エラー: {e}")
            # 一時ファイルが残っていたら削除
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise ExcelExportError(f"Excelバイナリデータの生成に失敗しました: {str(e)}") from e
    
    def _create_backup(self, file_path: Path) -> Path:
        """
        ファイルのバックアップを作成します。
        
        Args:
            file_path: バックアップするファイルのパス
            
        Returns:
            バックアップファイルのパス
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_{timestamp}_backup{file_path.suffix}")
        
        # ファイルをコピー
        import shutil
        shutil.copy2(file_path, backup_path)
        
        return backup_path
    
    def _set_headers(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        シートにヘッダーを設定します。
        
        Args:
            sheet: 設定対象のワークシート
        """
        headers = self.config.headers
        
        for col_letter, header_text in headers.items():
            cell = sheet[f"{col_letter}1"]
            cell.value = header_text
    
    def _add_data(self, sheet: openpyxl.worksheet.worksheet.Worksheet, results: List[ProcessResultProtocol]) -> None:
        """
        シートにデータを追加します。
        
        Args:
            sheet: 追加対象のワークシート
            results: 処理結果のリスト
        """
        for i, result in enumerate(results, start=2):  # ヘッダー行の次から
            try:
                # 辞書型かオブジェクト型かを判定
                if isinstance(result, dict):
                    # 辞書型の場合
                    # A列: スタイリスト名
                    stylist = result.get('selected_stylist', {})
                    sheet[f"A{i}"] = stylist.get('name', '') if isinstance(stylist, dict) else getattr(stylist, 'name', '')
                    
                    # B列: クーポン名
                    coupon = result.get('selected_coupon', {})
                    sheet[f"B{i}"] = coupon.get('name', '') if isinstance(coupon, dict) else getattr(coupon, 'name', '')
                    
                    # ユーザーが選択したテンプレートがある場合はそれを使用
                    template = result.get('user_selected_template', result.get('selected_template', {}))
                    
                    # C列: コメント
                    sheet[f"C{i}"] = template.get('comment', '') if isinstance(template, dict) else getattr(template, 'comment', '')
                    
                    # D列: スタイルタイトル
                    sheet[f"D{i}"] = template.get('title', '') if isinstance(template, dict) else getattr(template, 'title', '')
                else:
                    # オブジェクト型の場合
                    # A列: スタイリスト名
                    sheet[f"A{i}"] = getattr(result.selected_stylist, 'name', '')
                    
                    # B列: クーポン名
                    sheet[f"B{i}"] = getattr(result.selected_coupon, 'name', '')
                    
                    # ユーザーが選択したテンプレートがある場合はそれを使用
                    template_to_use = result.user_selected_template if hasattr(result, 'user_selected_template') and result.user_selected_template else result.selected_template
                    
                    # C列: コメント
                    sheet[f"C{i}"] = getattr(template_to_use, 'comment', '')
                    
                    # D列: スタイルタイトル
                    sheet[f"D{i}"] = getattr(template_to_use, 'title', '')
            except Exception as e:
                self.logger.error(f"Excelデータ追加エラー (行 {i}): {e}")
                # エラーが発生しても続行するために空の値を設定
                sheet[f"A{i}"] = "ERROR"
                sheet[f"B{i}"] = "ERROR"
                sheet[f"C{i}"] = "ERROR"
                sheet[f"D{i}"] = "ERROR"
            
            try:
                # E列: 性別
                if isinstance(result, dict):
                    attr_analysis = result.get('attribute_analysis', {})
                    sheet[f"E{i}"] = attr_analysis.get('sex', '') if isinstance(attr_analysis, dict) else getattr(attr_analysis, 'sex', '')
                else:
                    sheet[f"E{i}"] = getattr(result.attribute_analysis, 'sex', '')
                
                # F列: 長さ
                if isinstance(result, dict):
                    attr_analysis = result.get('attribute_analysis', {})
                    sheet[f"F{i}"] = attr_analysis.get('length', '') if isinstance(attr_analysis, dict) else getattr(attr_analysis, 'length', '')
                else:
                    sheet[f"F{i}"] = getattr(result.attribute_analysis, 'length', '')
                
                # G列: スタイルメニュー
                if isinstance(result, dict):
                    template = result.get('selected_template', {})
                    sheet[f"G{i}"] = template.get('menu', '') if isinstance(template, dict) else getattr(template, 'menu', '')
                else:
                    sheet[f"G{i}"] = getattr(result.selected_template, 'menu', '')
                
                # H列: ハッシュタグ
                if isinstance(result, dict):
                    template = result.get('selected_template', {})
                    hashtag_str = template.get('hashtag', '') if isinstance(template, dict) else getattr(template, 'hashtag', '')
                    # ハッシュタグを改行して表示し、最初の5個までに制限
                    hashtags = [tag.strip() for tag in hashtag_str.split(',') if tag.strip()][:5]
                    sheet[f"H{i}"] = '\n'.join(hashtags)
                    # セルの書式設定で改行を有効にする
                    sheet[f"H{i}"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    # 行の高さを調整（ハッシュタグの数に応じて）
                    row_height = min(15 * len(hashtags), 75)  # 1タグあたり15、最大75
                    sheet.row_dimensions[i].height = max(sheet.row_dimensions[i].height or 15, row_height)
                else:
                    hashtag_str = getattr(result.selected_template, 'hashtag', '')
                    # ハッシュタグを改行して表示し、最初の5個までに制限
                    hashtags = [tag.strip() for tag in hashtag_str.split(',') if tag.strip()][:5]
                    sheet[f"H{i}"] = '\n'.join(hashtags)
                    # セルの書式設定で改行を有効にする
                    sheet[f"H{i}"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    # 行の高さを調整（ハッシュタグの数に応じて）
                    row_height = min(15 * len(hashtags), 75)  # 1タグあたり15、最大75
                    sheet.row_dimensions[i].height = max(sheet.row_dimensions[i].height or 15, row_height)
                
                # 画像ファイル名の取得と変換
                if isinstance(result, dict):
                    image_name = result.get('image_name', '')
                else:
                    image_name = getattr(result, 'image_name', '')

                # ファイル名の変換（小文字で比較）
                image_name_lower = image_name.lower()
                if image_name_lower in self.filename_mapping:
                    image_name = self.filename_mapping[image_name_lower]
                    self.logger.info(f"ファイル名を変換: {image_name_lower} -> {image_name}")

                # I列: 画像ファイル名
                sheet[f"I{i}"] = image_name
            except Exception as e:
                self.logger.error(f"Excelデータ追加エラー2 (行 {i}): {e}")
                # エラーが発生しても続行するために空の値を設定
                sheet[f"E{i}"] = "ERROR"
                sheet[f"F{i}"] = "ERROR"
                sheet[f"G{i}"] = "ERROR"
                sheet[f"H{i}"] = "ERROR"
                sheet[f"I{i}"] = "ERROR"
    
    def _adjust_column_widths(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        列幅を自動調整します。
        
        Args:
            sheet: 調整対象のワークシート
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # ヘッダーセルの幅に対応するために係数を調整
            adjusted_width = (max_length + 2) * 1.1
            
            # 最大幅と最小幅を設定
            min_width = 10
            max_width = 50
            
            # ハッシュタグ列（H列）の場合は、最長のハッシュタグの長さに基づいて幅を設定
            if column_letter == 'H':
                # ハッシュタグは改行されるので、各行の最長のハッシュタグの長さを考慮
                max_tag_length = 0
                for cell in column:
                    if cell.value:
                        tags = str(cell.value).split('\n')
                        for tag in tags:
                            if len(tag) > max_tag_length:
                                max_tag_length = len(tag)
                
                # ハッシュタグ列の幅を調整
                adjusted_width = (max_tag_length + 2) * 1.1
                min_width = 15  # ハッシュタグ列の最小幅を大きめに
            
            # 幅を制限
            adjusted_width = max(min_width, min(adjusted_width, max_width))
            
            # 列幅を設定
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # _apply_styles メソッドを削除
